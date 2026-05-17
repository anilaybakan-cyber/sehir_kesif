#!/usr/bin/env python3
"""
assets/cities/*.json içindeki TÜM şehir paketlerini Excel'e döker
(cities_list ile sınırlı değil; bari, catania, batch dosyaları dahil).

Çıktı: ~/Desktop/içerikler1.xlsx

Her satırda benzersiz anahtar: kaynak_json (dosya adı .json olmadan).
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl gerekli: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
CITIES_DIR = ROOT / "assets" / "cities"
DESKTOP = Path.home() / "Desktop"
OUT_FILE = DESKTOP / "içerikler1.xlsx"


def S(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, (list, dict)):
        return json.dumps(v, ensure_ascii=False)
    return str(v).strip()


def load_json(path: Path) -> tuple[dict | None, str | None]:
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f), None
    except OSError as e:
        return None, str(e)
    except json.JSONDecodeError as e:
        return None, f"JSON hatası: {e.msg} (satır ~{e.lineno})"


def infer_city_from_stem(stem: str) -> tuple[str, str]:
    """batch / unique parça dosyaları için görünen şehir adı tahmini."""
    if "_unique" in stem:
        base = stem.split("_unique")[0]
    elif "_batch" in stem:
        base = stem.split("_batch")[0]
    else:
        base = stem
    name = base.replace("_", " ").strip().title()
    return name, name


def main() -> None:
    if not CITIES_DIR.is_dir():
        print(f"Klasör yok: {CITIES_DIR}", file=sys.stderr)
        sys.exit(1)

    json_paths = sorted(CITIES_DIR.glob("*.json"), key=lambda p: p.stem.lower())

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    ws_places = wb.create_sheet("Mekanlar", 0)
    ws_city = wb.create_sheet("Şehirler", 1)
    ws_guide = wb.create_sheet("Şehir_rehberi", 2)
    ws_routes = wb.create_sheet("Kürasyonlu_rotalar", 3)
    ws_food = wb.create_sheet("Yemek_rehberi", 4)

    hdr_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    places_headers = [
        "kaynak_json",
        "şehir_id",
        "şehir_adı_TR",
        "şehir_adı_EN",
        "ülke_TR",
        "ülke_EN",
        "yer_adı_TR",
        "yer_adı_EN",
        "kategori",
        "bölge_TR",
        "bölge_EN",
        "hakkında_TR",
        "hakkında_EN",
        "lokal_ipucu_TR",
        "lokal_ipucu_EN",
        "place_id",
    ]
    ws_places.append(places_headers)
    for c in range(1, len(places_headers) + 1):
        ws_places.cell(1, c).font = hdr_font

    city_headers = [
        "kaynak_json",
        "şehir_id",
        "şehir_adı_TR",
        "şehir_adı_EN",
        "ülke_TR",
        "ülke_EN",
        "kısa_açıklama_TR",
        "kısa_açıklama_EN",
        "şehir_localTips_liste_TR",
        "rehber_ipuçları_bölümü_TR",
        "rehber_ipuçları_bölümü_EN",
        "not",
    ]
    ws_city.append(city_headers)
    for c in range(1, len(city_headers) + 1):
        ws_city.cell(1, c).font = hdr_font

    guide_headers = [
        "kaynak_json",
        "şehir_id",
        "giriş_TR",
        "giriş_EN",
        "öneriler_TR",
        "öneriler_EN",
        "ipuçları_bölümü_TR",
        "ipuçları_bölümü_EN",
        "gizli_hazineler_TR",
        "gizli_hazineler_EN",
        "ulaşım_TR",
        "ulaşım_EN",
    ]
    ws_guide.append(guide_headers)
    for c in range(1, len(guide_headers) + 1):
        ws_guide.cell(1, c).font = hdr_font

    route_headers = [
        "kaynak_json",
        "şehir_id",
        "rota_id",
        "başlık_TR",
        "başlık_EN",
        "açıklama_TR",
        "açıklama_EN",
        "durak_idleri",
    ]
    ws_routes.append(route_headers)
    for c in range(1, len(route_headers) + 1):
        ws_routes.cell(1, c).font = hdr_font

    food_headers = [
        "kaynak_json",
        "şehir_id",
        "mutlaka_dene_TR",
        "yerel_içecekler_TR",
        "bahşiş_notu_TR",
    ]
    ws_food.append(food_headers)
    for c in range(1, len(food_headers) + 1):
        ws_food.cell(1, c).font = hdr_font

    bad_files: list[tuple[str, str]] = []
    ok_files = 0
    total_highlights = 0

    not_cell = (
        "şehir_id = kaynak_json (dosya kök adı). "
        "Aynı şehir için birden fazla dosya varsa (ör. catania.json + catania_batch_*) "
        "satırlar tekrarlanır; Excel'de kaynak_json ile süzebilirsiniz."
    )

    for path in json_paths:
        stem = path.stem
        cid = stem
        data, err = load_json(path)
        if data is None:
            bad_files.append((stem, err or "okunamadı"))
            continue

        # Kökü liste = yalnızca highlights parçası (catania_batch_*, bari_unique_*)
        if isinstance(data, list):
            ok_files += 1
            guess_tr, guess_en = infer_city_from_stem(stem)
            piece_note = (
                "Parça dosya: kök yapı highlights dizisi; şehir özeti/rehber/rota yok. "
                "Tam içerik için aynı şehrin ana .json dosyasına bakın (ör. catania.json)."
            )
            ws_city.append(
                [
                    stem,
                    cid,
                    guess_tr,
                    guess_en,
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    piece_note,
                ]
            )
            ws_guide.append([stem, cid] + [""] * (len(guide_headers) - 2))
            ws_food.append([stem, cid, "", "", ""])

            for h in data:
                if not isinstance(h, dict):
                    continue
                total_highlights += 1
                ws_places.append(
                    [
                        stem,
                        cid,
                        guess_tr,
                        guess_en,
                        "",
                        "",
                        S(h.get("name")),
                        S(h.get("name_en")),
                        S(h.get("category")),
                        S(h.get("area")),
                        S(h.get("area_en")),
                        S(h.get("description")),
                        S(h.get("description_en")),
                        S(h.get("tips")),
                        S(h.get("tips_en")),
                        S(h.get("id")),
                    ]
                )
            continue

        if not isinstance(data, dict):
            bad_files.append((stem, f"Beklenmeyen kök tip: {type(data).__name__}"))
            continue

        ok_files += 1
        city_json_name = S(data.get("city"))
        city_json_en = S(data.get("city_en"))
        desc_tr = S(data.get("description"))
        desc_en = S(data.get("description_en"))
        country_j_tr = S(data.get("country"))
        country_j_en = S(data.get("country_en"))

        tips_root = data.get("localTips") or []
        if isinstance(tips_root, list):
            tips_tr_joined = "\n".join(S(x) for x in tips_root)
        else:
            tips_tr_joined = S(tips_root)

        guide = data.get("guide")
        gtips_tr = S(guide.get("tips")) if isinstance(guide, dict) else ""
        gtips_en = S(guide.get("tips_en")) if isinstance(guide, dict) else ""

        ws_city.append(
            [
                stem,
                cid,
                city_json_name,
                city_json_en,
                country_j_tr,
                country_j_en,
                desc_tr,
                desc_en,
                tips_tr_joined,
                gtips_tr,
                gtips_en,
                not_cell,
            ]
        )

        if isinstance(guide, dict):
            ws_guide.append(
                [
                    stem,
                    cid,
                    S(guide.get("intro")),
                    S(guide.get("intro_en")),
                    S(guide.get("recommendations")),
                    S(guide.get("recommendations_en")),
                    S(guide.get("tips")),
                    S(guide.get("tips_en")),
                    S(guide.get("hidden_gems")),
                    S(guide.get("hidden_gems_en")),
                    S(guide.get("transport_guide")),
                    S(guide.get("transport_guide_en")),
                ]
            )
        else:
            ws_guide.append([stem, cid] + [""] * (len(guide_headers) - 2))

        fg = data.get("foodGuide")
        if isinstance(fg, dict):
            must_try = fg.get("must_try") or []
            drinks = fg.get("local_drinks") or []
            if isinstance(must_try, list):
                must_try_s = "\n".join(S(x) for x in must_try)
            else:
                must_try_s = S(must_try)
            if isinstance(drinks, list):
                drinks_s = "\n".join(S(x) for x in drinks)
            else:
                drinks_s = S(drinks)
            ws_food.append(
                [
                    stem,
                    cid,
                    must_try_s,
                    drinks_s,
                    S(fg.get("tipping")),
                ]
            )
        else:
            ws_food.append([stem, cid, "", "", ""])

        for route in data.get("curated_routes") or []:
            if not isinstance(route, dict):
                continue
            places_r = route.get("places") or []
            if isinstance(places_r, list):
                places_s = ", ".join(S(x) for x in places_r)
            else:
                places_s = S(places_r)
            ws_routes.append(
                [
                    stem,
                    cid,
                    S(route.get("id")),
                    S(route.get("title")),
                    S(route.get("title_en")),
                    S(route.get("description")),
                    S(route.get("description_en")),
                    places_s,
                ]
            )

        for h in data.get("highlights") or []:
            if not isinstance(h, dict):
                continue
            total_highlights += 1
            ws_places.append(
                [
                    stem,
                    cid,
                    city_json_name,
                    city_json_en,
                    country_j_tr,
                    country_j_en,
                    S(h.get("name")),
                    S(h.get("name_en")),
                    S(h.get("category")),
                    S(h.get("area")),
                    S(h.get("area_en")),
                    S(h.get("description")),
                    S(h.get("description_en")),
                    S(h.get("tips")),
                    S(h.get("tips_en")),
                    S(h.get("id")),
                ]
            )

    for ws in (ws_places, ws_city, ws_guide, ws_routes, ws_food):
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap

    def autosize(ws, max_w=60, max_rows=8000):
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = 10
            last = min(ws.max_row + 1, max_rows)
            for row in range(1, last):
                v = ws.cell(row, col).value
                if v:
                    line0 = str(v).split("\n")[0][:200]
                    max_len = min(max_w, max(max_len, len(line0)))
            ws.column_dimensions[letter].width = max_len + 2

    for ws in (ws_places, ws_city, ws_guide, ws_routes, ws_food):
        autosize(ws)

    ws_places.freeze_panes = "B2"
    ws_city.freeze_panes = "B2"

    summary = wb.create_sheet("_özet", 5)
    summary.append(["Toplam .json dosyası", len(json_paths)])
    summary.append(["Başarıyla işlenen", ok_files])
    summary.append(["Okunamayan / bozuk", len(bad_files)])
    summary.append(["Toplam mekan satırı", total_highlights])
    summary.append([])
    summary.append(["kaynak_json = assets/cities içindeki dosya adı (.json hariç)."])
    summary.append(["Tek şehir birden fazla dosyada olabilir (batch / unique); satırlar birleştirilmez."])

    if bad_files:
        note = wb.create_sheet("_hatalı_json", 6)
        note.append(["kaynak_json", "neden"])
        for stem, reason in bad_files:
            note.append([stem, reason])

    DESKTOP.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_FILE)
    print(f"Kaydedildi: {OUT_FILE}")
    print(f"İşlenen dosya: {ok_files} / {len(json_paths)}")
    print(f"Mekan satırı: {total_highlights}")
    if bad_files:
        print(f"Hatalı dosya: {len(bad_files)} → {', '.join(x[0] for x in bad_files)}")


if __name__ == "__main__":
    main()
