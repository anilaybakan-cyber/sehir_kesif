#!/usr/bin/env python3
import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def load_city_data(json_path):
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def translate_city(city_tr):
    city_translations = {
        "Atina": "Athens", "Belgrad": "Belgrade", "Brüksel": "Brussels", "Budapeşte": "Budapest",
        "Cenevre": "Geneva", "Fes": "Fez", "Floransa": "Florence", "İstanbul": "Istanbul",
        "Kahire": "Cairo", "Kapadokya": "Cappadocia", "Kopenhag": "Copenhagen", "Lizbon": "Lisbon",
        "Londra": "London", "Marakeş": "Marrakech", "Marsilya": "Marseille", "Midilli": "Lesbos",
        "Napoli": "Naples", "Paris": "Paris", "Prag": "Prague", "Roma": "Rome",
        "Saraybosna": "Sarajevo", "Seul": "Seoul", "Sevilla": "Seville", "Singapur": "Singapore",
        "Strazburg": "Strasbourg", "Viyana": "Vienna", "Zürih": "Zurich"
    }
    return city_translations.get(city_tr, city_tr)

def translate_category(category_tr):
    translations = {
        "Tarihi": "Historical", "Müze": "Museum", "Park": "Park", "Manzara": "Viewpoint",
        "Restoran": "Restaurant", "Kafe": "Cafe", "Bar": "Bar", "Alışveriş": "Shopping",
        "Deneyim": "Experience", "Sanat": "Art", "Mimari": "Architecture", "Doğa": "Nature",
        "Plaj": "Beach", "Eğlence": "Entertainment", "Spor": "Sports", "Dini": "Religious",
        "Yemek": "Food", "Gece Hayatı": "Nightlife", "Mahalle": "Neighborhood", "Sokak": "Street",
        "Köprü": "Bridge", "Kule": "Tower", "Saray": "Palace", "Kilise": "Church",
        "Cami": "Mosque", "Sinagog": "Synagogue", "Bahçe": "Garden", "Anıt": "Monument",
        "Heykel": "Statue", "Çeşme": "Fountain", "Meydan": "Square"
    }
    return translations.get(category_tr, category_tr)

def extract_data(source_dir):
    all_data = []
    source_path = Path(source_dir)
    for json_file in sorted(source_path.glob("*.json")):
        try:
            data = load_city_data(json_file)
            city_name = data.get("city", json_file.stem)
            highlights = data.get("highlights", [])
            for h in highlights:
                all_data.append({
                    "Şehir (TR)": city_name,
                    "Şehir (EN)": translate_city(city_name),
                    "Başlık (TR)": h.get("name", ""),
                    "Başlık (EN)": h.get("name_en", ""),
                    "Kategori (TR)": h.get("category", ""),
                    "Alt Kategori (TR)": h.get("subcategory", ""),
                    "Açıklama (TR)": h.get("description", ""),
                    "Açıklama (EN)": h.get("description_en", ""),
                    "Bölge": h.get("area", ""),
                    "İpuçları (TR)": h.get("tips", ""),
                    "İpuçları (EN)": h.get("tips_en", "")
                })
        except Exception as e:
            print(f"Error {json_file}: {e}")
    return all_data

def create_excel(data, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = list(data[0].keys()) if data else []
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font, cell.fill, cell.alignment = header_font, header_fill, alignment
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    ws.freeze_panes = "A2"
    wb.save(output_path)

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: script.py <source_dir> <output_xlsx>")
        sys.exit(1)
    
    source = sys.argv[1]
    output = sys.argv[2]
    data = extract_data(source)
    if data:
        create_excel(data, output)
        print(f"Success: {output}")
    else:
        print("No data found.")
