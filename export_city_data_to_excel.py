#!/usr/bin/env python3
"""
Tüm şehir JSON dosyalarından başlık, şehir, içerik ve kategori bilgilerini
hem Türkçe hem İngilizce olarak Excel'e aktarır.
"""

import json
import os
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("openpyxl modülü yükleniyor...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Paths
CITIES_DIR = Path(__file__).parent / "ota_data_pack" / "cities"
OUTPUT_FILE = Path(__file__).parent / "uygulama_verileri_iki_dil.xlsx"

def load_city_data(json_path):
    """Load data from a city JSON file."""
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def extract_all_data():
    """Extract all data from all city files."""
    all_data = []
    
    for json_file in sorted(CITIES_DIR.glob("*.json")):
        try:
            data = load_city_data(json_file)
            city_name = data.get("city", json_file.stem)
            country = data.get("country", "")
            city_description = data.get("description", "")
            
            # Get highlights (places)
            highlights = data.get("highlights", [])
            
            for highlight in highlights:
                row = {
                    "Şehir (TR)": city_name,
                    "Şehir (EN)": translate_city(city_name),
                    "Ülke": country,
                    "Başlık (TR)": highlight.get("name", ""),
                    "Başlık (EN)": highlight.get("name_en", highlight.get("name", "")),
                    "Kategori (TR)": highlight.get("category", ""),
                    "Kategori (EN)": translate_category(highlight.get("category", "")),
                    "Alt Kategori (TR)": highlight.get("subcategory", highlight.get("category", "")),
                    "Alt Kategori (EN)": translate_category(highlight.get("subcategory", highlight.get("category", ""))),
                    "Açıklama (TR)": highlight.get("description", ""),
                    "Açıklama (EN)": highlight.get("description_en", ""),
                    "Bölge": highlight.get("area", ""),
                    "İpuçları (TR)": highlight.get("tips", ""),
                    "İpuçları (EN)": highlight.get("tips_en", ""),
                }
                all_data.append(row)
                
        except Exception as e:
            print(f"Error processing {json_file}: {e}")
    
    return all_data

def translate_city(city_tr):
    """Translate Turkish city name to English."""
    city_translations = {
        "Amsterdam": "Amsterdam",
        "Antalya": "Antalya",
        "Atina": "Athens",
        "Bangkok": "Bangkok",
        "Barcelona": "Barcelona",
        "Belgrad": "Belgrade",
        "Berlin": "Berlin",
        "Bologna": "Bologna",
        "Brugge": "Bruges",
        "Brüksel": "Brussels",
        "Budapeşte": "Budapest",
        "Cenevre": "Geneva",
        "Colmar": "Colmar",
        "Dubai": "Dubai",
        "Dublin": "Dublin",
        "Edinburgh": "Edinburgh",
        "Fes": "Fez",
        "Floransa": "Florence",
        "Gaziantep": "Gaziantep",
        "Giethoorn": "Giethoorn",
        "Hallstatt": "Hallstatt",
        "Heidelberg": "Heidelberg",
        "Hong Kong": "Hong Kong",
        "İstanbul": "Istanbul",
        "Kahire": "Cairo",
        "Kapadokya": "Cappadocia",
        "Kopenhag": "Copenhagen",
        "Kotor": "Kotor",
        "Lizbon": "Lisbon",
        "Londra": "London",
        "Lucerne": "Lucerne",
        "Lyon": "Lyon",
        "Madrid": "Madrid",
        "Marakeş": "Marrakech",
        "Marsilya": "Marseille",
        "Matera": "Matera",
        "Midilli": "Lesbos",
        "Milano": "Milan",
        "Napoli": "Naples",
        "New York": "New York",
        "Nice": "Nice",
        "Oslo": "Oslo",
        "Paris": "Paris",
        "Porto": "Porto",
        "Prag": "Prague",
        "Roma": "Rome",
        "Rovaniemi": "Rovaniemi",
        "San Sebastian": "San Sebastian",
        "Santorini": "Santorini",
        "Saraybosna": "Sarajevo",
        "Seul": "Seoul",
        "Sevilla": "Seville",
        "Singapur": "Singapore",
        "Sintra": "Sintra",
        "Stockholm": "Stockholm",
        "Strazburg": "Strasbourg",
        "Tokyo": "Tokyo",
        "Tromsø": "Tromsø",
        "Venedik": "Venice",
        "Viyana": "Vienna",
        "Zermatt": "Zermatt",
        "Zürih": "Zurich",
    }
    return city_translations.get(city_tr, city_tr)

def translate_category(category_tr):
    """Translate Turkish category to English."""
    translations = {
        "Tarihi": "Historical",
        "Müze": "Museum",
        "Park": "Park",
        "Manzara": "Viewpoint",
        "Restoran": "Restaurant",
        "Kafe": "Cafe",
        "Bar": "Bar",
        "Alışveriş": "Shopping",
        "Deneyim": "Experience",
        "Sanat": "Art",
        "Mimari": "Architecture",
        "Doğa": "Nature",
        "Plaj": "Beach",
        "Eğlence": "Entertainment",
        "Spor": "Sports",
        "Dini": "Religious",
        "Yemek": "Food",
        "Gece Hayatı": "Nightlife",
        "Mahalle": "Neighborhood",
        "Sokak": "Street",
        "Köprü": "Bridge",
        "Kule": "Tower",
        "Saray": "Palace",
        "Kilise": "Church",
        "Cami": "Mosque",
        "Sinagog": "Synagogue",
        "Bahçe": "Garden",
        "Anıt": "Monument",
        "Heykel": "Statue",
        "Çeşme": "Fountain",
        "Meydanlar": "Squares",
        "Meydan": "Square",
    }
    return translations.get(category_tr, category_tr)

def create_excel(data):
    """Create Excel file with the data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Şehir Verileri"
    
    # Headers
    headers = list(data[0].keys()) if data else []
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border
    
    # Write data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border
    
    # Adjust column widths
    column_widths = {
        "Şehir (TR)": 15,
        "Şehir (EN)": 15,
        "Ülke": 15,
        "Başlık (TR)": 35,
        "Başlık (EN)": 35,
        "Kategori (TR)": 15,
        "Kategori (EN)": 15,
        "Alt Kategori (TR)": 15,
        "Alt Kategori (EN)": 15,
        "Açıklama (TR)": 50,
        "Açıklama (EN)": 50,
        "Bölge": 20,
        "İpuçları (TR)": 40,
        "İpuçları (EN)": 40,
    }
    
    for col, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = column_widths.get(header, 20)
    
    # Freeze first row
    ws.freeze_panes = "A2"
    
    # Save
    wb.save(OUTPUT_FILE)
    print(f"Excel dosyası oluşturuldu: {OUTPUT_FILE}")
    return OUTPUT_FILE

def main():
    print("Şehir verilerini topluyorum...")
    all_data = extract_all_data()
    print(f"Toplam {len(all_data)} kayıt bulundu.")
    
    if all_data:
        output_path = create_excel(all_data)
        print(f"\n✓ Excel dosyası başarıyla oluşturuldu!")
        print(f"  Dosya: {output_path}")
        print(f"  Toplam kayıt: {len(all_data)}")
    else:
        print("Hiç veri bulunamadı!")

if __name__ == "__main__":
    main()
